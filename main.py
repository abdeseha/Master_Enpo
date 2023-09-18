from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QCheckBox , QLineEdit, QWidget, QComboBox
from PyQt5.QtWidgets import QSpinBox, QVBoxLayout, QFileDialog, QAction,QTableWidget, QTableWidgetItem
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5 import uic
import threading
import openpyxl
import matplotlib.pyplot as plt
import sys
from analyse import Analyse_C
from plan import Plan

class App(QMainWindow):
    def __init__(self):
        super(App, self).__init__()

        self.files = None
        self.opt_data = {}
        self.opt_id = ""
        self.type_list_table = []
        self.oprts_table = []
        self.result_tab = []
        self.total_time = []
        self.total_power = []

        uic.loadUi("./assets/app_gui.ui", self)

        # -------------------- toolbare -------------------------#
        self.new_file = self.findChild(QAction, "actionNew_File")
        self.save_tab = self.findChild(QAction, "actionSave_Table")
        self.save_points = self.findChild(QAction, "actionSave_points")
        self.save_sched = self.findChild(QAction, "actionSchedule")

        #-----------------------tabs------------------------#
        self.ana = self.findChild(QWidget, 'analyse')
        self.plan = self.findChild(QWidget, 'plan')
        self.settings = self.findChild(QWidget, 'settings')

        #-------------------------graph-------------------------#
        self.lay = self.findChild(QVBoxLayout, "plot_holder")
        self.figure = plt.Figure()
        self.canvas = FigureCanvas(self.figure)
        self.lay.addWidget(self.canvas)

        #------------------------Analyse Buttons----------------------------#
        self.apply = self.findChild(QPushButton, "apply")
        self.apply.setEnabled(False)
        self.stop = self.findChild(QPushButton, "stop")
        self.stop.setEnabled(False)
        self.chfile = self.findChild(QPushButton, "file")
        self.to_pause = self.findChild(QPushButton, "to_pause")
        self.to_pause.setEnabled(False)
        self.to_continue = self.findChild(QPushButton, "to_continue")
        self.to_continue.setEnabled(False)

        #--------------------------Plan buttons--------------------------#
        self.cr_sched = self.findChild(QPushButton, "cr_sched")
        self.start_sched = self.findChild(QPushButton, "start_sched")
        self.start_sched.setEnabled(False)
        self.add_mach = self.findChild(QPushButton, "add_mach")
        self.add_mach.setEnabled(False)
        self.rm_mach = self.findChild(QPushButton, "rm_mach")
        self.rm_mach.setEnabled(False)
        self.add_oprt = self.findChild(QPushButton, "add_oprt")
        self.add_oprt.setEnabled(False)
        self.rm_oprt = self.findChild(QPushButton, "rm_oprt")
        self.rm_oprt.setEnabled(False)

        #--------------------Plan spinbox------------------------#
        self.num_mach_type = self.findChild(QSpinBox, "num_mach_type")
        self.num_mach_type.setEnabled(False)
        self.max_pow = self.findChild(QSpinBox, "max_pow")
        self.max_pow.setEnabled(False)

        #--------------------Plan table---------------------#
        self.tab_plan = self.findChild(QTableWidget, 'tab_plan')
        self.make_type_list()

        #----------------------------settings buttons------------------------#
        self.crt_new_opt = self.findChild(QPushButton, "crt_new_opt")
        self.mod_opt = self.findChild(QPushButton, "mod_opt")
        self.new_line = self.findChild(QPushButton, "new_line")
        self.new_line.setEnabled(False)
        self.apply_new_opt = self.findChild(QPushButton, "apply_new_opt")
        self.apply_new_opt.setEnabled(False)
        self.remove_opt = self.findChild(QPushButton, "remove_opt")
        self.remove_opt.setEnabled(False) 

        #----------------------------settings tables---------------------#
        self.add_opt = self.findChild(QTableWidget, 'add_opt')
        self.add_opt.setEnabled(False)
        self.opts = self.findChild(QTableWidget, 'opts')
        self.update_opt_tab()
        self.opts.setEnabled(False)

        #--------------------------settings Editline-------------------------#
        self.opt_name = self.findChild(QLineEdit, 'opt_name')
        self.opt_name.setEnabled(False)

        #-------------------------settings spinbox------------------------#
        self.skip_row_box = self.findChild(QSpinBox, 'skip_row_box')
        self.rows_box = self.findChild(QSpinBox, 'rows_box')
        self.num_vals = self.findChild(QSpinBox, "num_vals")
        self.num_times = self.findChild(QSpinBox, "num_times")
        
        #--------------------------settings checkBox------------------------#
        self.mob_pressure = self.findChild(QCheckBox, 'Mob_pressure')
        self.tot_pow = self.findChild(QCheckBox, 'tot_pow')
        
        app.aboutToQuit.connect(self.stop_app)
        
        #----------------------tool bare--------------------------#
        self.new_file.triggered.connect(self.creat_new_file)
        self.save_tab.triggered.connect(self.save_opts)
        self.save_points.triggered.connect(self.save_opts_points)
        self.save_sched.triggered.connect(self.saveing_sched)

        #------------------------Analyse---------------------#
        self.apply.clicked.connect(self.start_analyse)
        self.stop.clicked.connect(self.stop_analyse)
        self.chfile.clicked.connect(self.choose_file)
        self.to_pause.clicked.connect(self.pausing)
        self.to_continue.clicked.connect(self.continueing)
        
        #----------------------plan------------------#
        self.cr_sched.clicked.connect(self.crting_sched)
        self.add_oprt.clicked.connect(self.add_oprt_table)
        self.rm_oprt.clicked.connect(self.rm_opt_table)
        self.add_mach.clicked.connect(self.add_machine)
        self.rm_mach.clicked.connect(self.rm_machine)
        self.start_sched.clicked.connect(self.read_sched)
        self.num_mach_type.valueChanged.connect(self.update_type)

        #---------------------------settings---------------------#
        self.new_line.clicked.connect(self.add_row)
        self.crt_new_opt.clicked.connect(self.new_opt)
        self.mod_opt.clicked.connect(self.mod_opts)
        self.apply_new_opt.clicked.connect(self.read_opt)
        self.remove_opt.clicked.connect(self.rmv_opt)
        self.skip_row_box.valueChanged.connect(self.update_min)

        self.show()

    def choose_file(self):
        self.files, _ = QFileDialog.getOpenFileNames(self, "Open File", "", "Excel Files (*.xlsx *.csv)")
        if self.files != []:
            Analyse_C.new(len(self.files))

            self.stop.setEnabled(False)
            self.apply.setEnabled(True)
            self.to_continue.setEnabled(False)
            self.to_pause.setEnabled(False)
    
    def start_analyse(self):
        Analyse_C.on = True

        self.stop.setEnabled(True)
        self.to_pause.setEnabled(True)
        self.apply.setEnabled(False)
        self.chfile.setEnabled(False)
        self.to_continue.setEnabled(False)
        self.plan.setEnabled(False)
        self.settings.setEnabled(False)

        chart = Analyse_C(files=self.files,
                          row=self.rows_box.value(), skiprow=self.skip_row_box.value(),
                          marge=self.num_vals.value(), times=self.num_times.value(),
                          )
        analyse_thread = threading.Thread(target=chart.draw_chart, args = (self, self.mob_pressure.isChecked(), self.tot_pow.isChecked()))
        analyse_thread.start()
        
    def pausing(self):
        Analyse_C.on = False
        if not Analyse_C.in_loop:
            self.to_continue.setEnabled(True)
            self.to_pause.setEnabled(False)

    def continueing(self):
        Analyse_C.on = True
        self.start_analyse()

    def stop_analyse(self):
        Analyse_C.on = False
        if not Analyse_C.in_loop:
            self.stop.setEnabled(False)
            self.apply.setEnabled(False)
            self.chfile.setEnabled(True)
            self.to_continue.setEnabled(False)
            self.to_pause.setEnabled(False)
            self.settings.setEnabled(True)
            self.plan.setEnabled(True)

    def stop_app(self):
        Analyse_C.on = False
        Analyse_C.in_loop = False
        
    def update_min(self):
        self.rows_box.setMinimum(self.skip_row_box.value())

    def creat_new_file(self):
        file_name, _ = QFileDialog.getSaveFileName(caption="Make a new file", directory="", filter="CSV (*.csv);;Excel (*.xlsx)")
        try:
            if file_name.lower().endswith(".csv"):
                with open(file_name, 'w') as file:
                    file.write("t(s),I1 (A),I2 (A),I3 (A),V (V)")
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet["A1"] = 't(s)'
                sheet["B1"] = 'I1 (A)'
                sheet["C1"] = 'I2 (A)'
                sheet["D1"] = 'I3 (A)'
                sheet["E1"] = 'V (V)'
                workbook.save(file_name)
        except FileNotFoundError:
            return
        
    def save_opts(self):
        try:
            for file, opt_f in zip(self.files, Analyse_C.file_opts.values()):
                fl = file.split('/')[-1]
                file_name, _ = QFileDialog.getSaveFileName(caption=f"Save the results of {fl}", directory="", filter="Excel (*.xlsx)")
                try:
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    index = 1
                    for opt, charact in opt_f.items():
                        sheet["A"+str(index)] = f"{opt} :"
                        sheet["B"+str(index)] = "Number of use"
                        sheet["B"+str(index+1)] = charact['nb_of_use']
                        sheet["C"+str(index)] = "Stabal value"
                        sheet["C"+str(index+1)] = charact["stab_value"]
                        sheet["D"+str(index)] = "Start times"
                        sheet["E"+str(index)] = "End times"
                        sheet["F"+str(index)] = "Max value"
                        for i in range(len(charact['start_time'])):
                            index += 1
                            sheet["D"+str(index)] = charact['start_time'][i]
                            sheet["E"+str(index)] =  charact['end_time'][i]
                            sheet["F"+str(index)] = charact['max'][i]

                        index += 2
                    workbook.save(file_name)
                except (FileNotFoundError, PermissionError):
                    return
        except (TypeError, PermissionError):
            return
            
    def save_opts_points(self):
        try:
            new_op_fs = {}
            for file, opt_f in zip(self.files, Analyse_C.file_dump_ops.values()):
                new_op_f = []
                for op in opt_f:
                    new_op = {}
                    t_0 = next(iter(op))
                    for time, value in op.items():
                        new_op[time-t_0] = value
                    new_op_f.append(new_op)
                new_op_fs[file] = new_op_f
            
            for file, opt_f in zip(self.files, new_op_fs.values()):
                fl = file.split('/')[-1]
                file_name, _ = QFileDialog.getSaveFileName(caption=f"Save the results of {fl}", directory="", filter="Excel (*.xlsx)")
                
                try:
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    index = 1
                    for i, opt in enumerate(opt_f):
                        sheet["A"+str(index)] = f"operation_{i} :"
                        sheet["B"+str(index)] = "time :"
                        sheet["C"+str(index)] = "value :"
                        for t, val in opt.items():
                            index += 1
                            sheet["B"+str(index)] = t
                            sheet["C"+str(index)] = val

                        index += 2
                        workbook.save(file_name)
                        
                except (FileNotFoundError, PermissionError):
                    return
        except (TypeError, PermissionError):
            return
            
    def add_row(self):
        self.add_opt.insertRow(self.add_opt.rowCount())
        return
    
    def new_opt(self):
        self.add_opt.setEnabled(True)
        self.opt_name.setEnabled(True)
        self.new_line.setEnabled(True)
        self.apply_new_opt.setEnabled(True)
        return
    
    def mod_opts(self):
        self.opts.setEnabled(True)
        self.remove_opt.setEnabled(True)
        return
    
    def read_opt(self):
        rows = self.add_opt.rowCount()
        self.opt_id = self.opt_name.text()

        for row in range(rows):
            t = self.add_opt.item(row, 0)
            val = self.add_opt.item(row, 1)
            if t is not None and val is not None:
                try:
                    self.opt_data[float(t.text())] = float(val.text())
                except ValueError:
                    continue   

        self.opt_data = dict(sorted(self.opt_data.items()))
        
        self.opts_list = [op[0] for op in Plan().add_to_db(self.opt_data, self.opt_id)]
        self.update_opt_tab()
        
        self.add_opt.setRowCount(0)
        self.opt_name.setText('')
        self.opt_data = {}
        self.update_opts()
        return

    def update_opt_tab(self):
        self.opts_list = [op[0] for op in Plan().add_to_db({}, '')]
        self.opts.setRowCount(len(self.opts_list))
        for i, op in enumerate(self.opts_list):
            self.opts.setItem(i, 0, QTableWidgetItem(op))
        return

    def rmv_opt(self):
        selected_rows = set()
        for item in self.opts.selectedItems():
            selected_rows.add(item.row())

        for row_index in sorted(selected_rows):
            opt = self.opts.item(row_index, 0).text()
            Plan().rm_opt(opt)
            self.opts.removeRow(row_index)
        self.update_opts()
        return
        
    def crting_sched(self):
        self.cr_sched.setEnabled(False)
        self.tab_plan.setEnabled(True)
        self.start_sched.setEnabled(True)
        self.add_mach.setEnabled(True)
        self.rm_mach.setEnabled(True)
        self.add_oprt.setEnabled(True)
        self.rm_oprt.setEnabled(True)
        self.max_pow.setEnabled(True)
        self.num_mach_type.setEnabled(True)
        self.ana.setEnabled(True)
        self.settings.setEnabled(True)
        self.tab_plan.setRowCount(1)
        self.tab_plan.setColumnCount(1)
        self.tab_plan.setHorizontalHeaderLabels(['Machine 1'])
        self.type_list_table = []
        self.oprts_table = []
        self.make_type_list()

        return

    def rm_opt_table(self):
        row_index = self.tab_plan.currentRow()
        if row_index > 0:
            self.tab_plan.removeRow(row_index)
            self.oprts_table.pop(row_index-1)
        return

    def add_machine(self):
        column_count = self.tab_plan.columnCount()
        self.tab_plan.setColumnCount(column_count + 1)
        self.tab_plan.setHorizontalHeaderItem(column_count, QTableWidgetItem(f'Machine {column_count + 1}'))
        self.make_type_list()
        return
    
    def rm_machine(self):
        selected_column = self.tab_plan.currentColumn()

        if selected_column >= 0:
            self.tab_plan.removeColumn(selected_column)
            self.type_list_table.pop(selected_column)
            for row in range(self.tab_plan.rowCount() -1):
                self.oprts_table[row].pop(selected_column)
        
        column_count = self.tab_plan.columnCount()
        new_headers = []
        for cul in range(1, column_count+1):
            new_headers.append(f'Machine {cul}')

        self.tab_plan.setHorizontalHeaderLabels(new_headers)
        return
    
    def make_type_list(self):
        type_list = QComboBox()
        for i in range(1, self.num_mach_type.value()+1):
            type_list.addItem(f'Type {i}')
        
        column_count = self.tab_plan.columnCount()

        self.type_list_table.append(type_list)
        self.tab_plan.setCellWidget(0, column_count-1, self.type_list_table[column_count-1])

        row_count = self.tab_plan.rowCount()
        if row_count > 1:
            for row in range(row_count -1):
                oprt_list = QComboBox()
                oprt_list.addItem('NONE')
                for i in self.opts_list:
                    oprt_list.addItem(i)
                self.oprts_table[row].append(oprt_list)
                self.tab_plan.setCellWidget(row+1, column_count-1, self.oprts_table[row][column_count-1])
        return
    
    def update_type(self):
        for item in self.type_list_table:
            item.clear()
            for i in range(1, self.num_mach_type.value()+1):
                item.addItem(f'Type {i}')
        return

    def make_oprt_list(self):
        self.opts_list = [op[0] for op in Plan().add_to_db({}, '')]
        self.oprts_table.append([])
        row_count = self.tab_plan.rowCount()
        for index in range(self.tab_plan.columnCount()):
            oprt_list = QComboBox()
            oprt_list.addItem('NONE')
            oprt_list.addItems(self.opts_list)
            self.oprts_table[row_count-2].append(oprt_list)
            self.tab_plan.setCellWidget(row_count-1, index, self.oprts_table[row_count-2][index])
        return
    
    def add_oprt_table(self):
        self.tab_plan.insertRow(self.tab_plan.rowCount())
        self.make_oprt_list()
        return

    def update_opts(self):
        for x in self.oprts_table:
            for y in x:
                y.clear()
                y.addItem('None')
                self.opts_list = [op[0] for op in Plan().add_to_db({}, '')]
                for i in self.opts_list:
                    y.addItem(i)
        return
    
    def read_sched(self):
        rows = self.tab_plan.rowCount()
        cols = self.tab_plan.columnCount()
        self.table = []

        for col in range(cols):
            self.table.append([])
            for row in range(rows):
                item = self.tab_plan.cellWidget(row, col)
                if item is not None:
                    self.table[col].append(item.currentText())

        if len(self.table[0]) > 1:    
            self.start_sched.setEnabled(False)
            self.add_mach.setEnabled(False)
            self.rm_mach.setEnabled(False)
            self.add_oprt.setEnabled(False)
            self.rm_oprt.setEnabled(False)
            self.num_mach_type.setEnabled(False)
            self.max_pow.setEnabled(False)
            self.tab_plan.setEnabled(False)

            self.starting_sched()
            sched = threading.Thread(target=self.starting_sched())
            sched.start()
        
        return
    
    def saveing_sched(self):
        if self.result_tab != []:
            try:
                file_name, _ = QFileDialog.getSaveFileName(caption=f"Save the table", directory="", filter="Excel (*.xlsx)")    
                try:
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active

                    for i, mach in enumerate(self.table):
                        cell = sheet.cell(row=1, column=i+1, value= f'Machine {i+1}')
                        cell = sheet.cell(row=2 , column=i+1, value = mach[0])

                    for j, y in enumerate(self.result_tab):
                        for i,x in enumerate(y):
                            cell = sheet.cell(row = i+3, column=j+1, value = x)
                    
                    workbook.save(file_name)
                except FileNotFoundError:
                    return
                
                file_name, _ = QFileDialog.getSaveFileName(caption=f"Save the Electricity power", directory="", filter="Excel (*.xlsx)")
                try:
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    
                    cell = sheet.cell(row=1, column=1, value= 'Time (s)')
                    cell = sheet.cell(row=1, column=2, value= 'Power (W)')
                    for i, t in enumerate(self.total_time):
                        cell = sheet.cell(row=i+2, column=1, value= t)
                    for i, p in enumerate(self.total_power):
                        cell = sheet.cell(row=i+2, column=2, value= p)
                    workbook.save(file_name)
                except (FileNotFoundError, PermissionError):
                    return
            except (TypeError, PermissionError):
                return
        else:
            return

    def starting_sched(self):
        max_pow = self.max_pow.value()
        self.result_tab, self.total_time, self.total_power = Plan().schedule(self.table, max_pow, self)
        if self.result_tab != []:
            self.tab_plan.setRowCount(1)
            lin = 1
            for i in self.result_tab:
                if lin < len(i): lin = len(i)
            self.tab_plan.setRowCount(lin+1)
            for i in range(len(self.result_tab)):
                for j in range(1,len(self.result_tab[i])+1):
                    self.tab_plan.setItem(j, i, QTableWidgetItem(f"{self.result_tab[i][j-1]}"))
        return

app = QApplication(sys.argv)
UIWindow = App()
app.exec()